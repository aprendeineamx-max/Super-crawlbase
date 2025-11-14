"""Script de prueba para el sistema de scrapers."""
import json
import sys
from pathlib import Path

# Agregar el backend al path
sys.path.insert(0, str(Path(__file__).parent))

# Importar modelos primero para evitar problemas de relaciones
from app.profiles import models as profile_models
from app.projects import models as project_models

from app.core.database import Session, engine, init_db
from app.profiles.service import ProfileService
from app.scrapers import service

# Inicializar base de datos
init_db()

# Obtener el perfil demo
with Session(engine) as session:
    profile_service = ProfileService(session=session)
    profiles = list(profile_service.list(include_tokens=True))
    
    if not profiles:
        print("❌ No se encontraron perfiles")
        sys.exit(1)
    
    demo_profile = profiles[0]
    print(f"✅ Perfil encontrado: {demo_profile.name} (ID: {demo_profile.id})")
    
    # URLs de prueba
    test_urls = [
        "https://example.com",
        "https://httpbin.org/get",
    ]
    
    print(f"\n📋 Probando scraping de {len(test_urls)} URLs...")
    print(f"URLs: {test_urls}\n")
    
    try:
        # Probar scraping
        result = service.scrape_urls(
            urls=test_urls,
            profile=demo_profile,
            excel_path=None,  # Crear nuevo archivo
            sheet_name="Test_Scraping",
        )
        
        print("✅ Scraping completado exitosamente!")
        print(f"   - Total URLs: {result['total_urls']}")
        print(f"   - Exitosos: {result['successful']}")
        print(f"   - Fallidos: {result['failed']}")
        print(f"   - Archivo Excel: {result['excel_path']}")
        print(f"   - Hoja: {result['sheet_name']}")
        print(f"   - Mensaje: {result['message']}")
        
        if result['errors']:
            print(f"\n⚠️  Errores encontrados:")
            for error in result['errors']:
                print(f"   - {error['url']}: {error['error']}")
        
        # Verificar que el archivo existe
        excel_file = Path(result['excel_path'])
        if excel_file.exists():
            print(f"\n✅ Archivo Excel creado correctamente: {excel_file}")
            print(f"   Tamaño: {excel_file.stat().st_size} bytes")
        else:
            print(f"\n❌ El archivo Excel no se encontró en: {excel_file}")
        
        # Probar con archivo existente (agregar nueva hoja)
        print(f"\n📋 Probando agregar nueva hoja a archivo existente...")
        result2 = service.scrape_urls(
            urls=["https://example.com"],
            profile=demo_profile,
            excel_path=excel_file.name,  # Usar el archivo creado anteriormente
            sheet_name="Segunda_Hoja",
        )
        
        print("✅ Segunda hoja agregada exitosamente!")
        print(f"   - Archivo Excel: {result2['excel_path']}")
        print(f"   - Hoja: {result2['sheet_name']}")
        
        # Verificar que el archivo tiene múltiples hojas
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        print(f"\n✅ Verificación de hojas en Excel:")
        print(f"   - Total de hojas: {len(wb.sheetnames)}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"   - Hoja '{sheet_name}': {ws.max_row} filas, {ws.max_column} columnas")
        
        print("\n✅ Todas las pruebas completadas exitosamente!")
        
    except Exception as e:
        print(f"\n❌ Error durante las pruebas: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

