from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.crawlbase.client import CrawlbaseClient, CrawlbaseClientError
from app.profiles.schemas import ProfileReadSchema

logger = logging.getLogger(__name__)


def scrape_url(client: CrawlbaseClient, url: str) -> Dict[str, Any]:
    """Scrapea una URL usando Crawlbase API."""
    try:
        # Usar el endpoint de crawling API
        response = client.get("/", params={"url": url})
        
        return {
            "url": url,
            "success": response.status_code == 200,
            "status_code": response.status_code,
            "data": response.data,
            "headers": response.headers,
            "error": None,
        }
    except CrawlbaseClientError as e:
        logger.error(f"Error de Crawlbase al scrapear {url}: {e}")
        return {
            "url": url,
            "success": False,
            "status_code": None,
            "data": None,
            "headers": {},
            "error": str(e),
        }
    except Exception as e:
        logger.error(f"Error inesperado al scrapear {url}: {e}")
        return {
            "url": url,
            "success": False,
            "status_code": None,
            "data": None,
            "headers": {},
            "error": str(e),
        }


def _extract_data_from_response(scrape_result: Dict[str, Any]) -> Dict[str, Any]:
    """Extrae datos estructurados de la respuesta de Crawlbase."""
    data = scrape_result.get("data", {})
    
    # Información básica
    extracted = {
        "url": scrape_result.get("url", ""),
        "status_code": scrape_result.get("status_code", ""),
        "success": scrape_result.get("success", False),
        "error": scrape_result.get("error"),
    }
    
    # Extraer información adicional de headers si están disponibles
    headers = scrape_result.get("headers", {})
    if headers:
        extracted["content_type"] = headers.get("content-type", headers.get("Content-Type", ""))
        extracted["content_length"] = headers.get("content-length", headers.get("Content-Length", ""))
    
    # Procesar el cuerpo de la respuesta
    if isinstance(data, dict):
        if "body" in data:
            # Es HTML, extraer información básica
            body = str(data.get("body", ""))
            extracted["content_type"] = extracted.get("content_type", "text/html")
            extracted["content_length"] = len(body)
            # Intentar extraer título si es HTML
            if "<title>" in body:
                try:
                    title_start = body.find("<title>") + 7
                    title_end = body.find("</title>", title_start)
                    if title_end > title_start:
                        extracted["page_title"] = body[title_start:title_end].strip()[:200]
                except:
                    pass
        else:
            # Es JSON, agregar campos relevantes
            for key, value in data.items():
                if key not in extracted and not isinstance(value, (dict, list)):
                    extracted[key] = str(value)[:500]  # Limitar longitud
    elif isinstance(data, str):
        extracted["content"] = data[:500]  # Primeros 500 caracteres
        extracted["content_length"] = len(data)
    
    return extracted


def save_to_excel(
    results: List[Dict[str, Any]],
    excel_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> tuple[str, str]:
    """Guarda los resultados del scraping en un archivo Excel.
    
    Args:
        results: Lista de resultados del scraping
        excel_path: Ruta del archivo Excel (si existe, se agrega una hoja; si no, se crea)
        sheet_name: Nombre de la hoja (si no se proporciona, se genera automáticamente)
    
    Returns:
        Tupla con (ruta_absoluta_del_excel, nombre_de_la_hoja)
    """
    # Generar nombre de hoja si no se proporciona
    if not sheet_name:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sheet_name = f"Scraping_{timestamp}"
    
    # Limpiar nombre de hoja (Excel tiene límites)
    sheet_name = sheet_name[:31]  # Excel limita a 31 caracteres
    
    # Determinar ruta del archivo
    output_dir = Path(__file__).resolve().parents[2] / "output"
    output_dir.mkdir(exist_ok=True)
    
    if excel_path:
        excel_file = Path(excel_path)
        # Si es relativa o solo un nombre de archivo, usar el directorio de salida
        if not excel_file.is_absolute() or not excel_file.parent.exists():
            excel_file = output_dir / excel_path
        # Si el archivo no existe pero la ruta es absoluta y válida, usarla
        # Si no existe, crear en output_dir
        if not excel_file.exists() and excel_file.parent != output_dir:
            excel_file = output_dir / excel_file.name
    else:
        # Crear nuevo archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"scraping_{timestamp}.xlsx"
    
    # Cargar o crear workbook
    if excel_file.exists():
        wb = load_workbook(excel_file)
        # Verificar si la hoja ya existe
        if sheet_name in wb.sheetnames:
            # Agregar sufijo numérico
            counter = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1
    else:
        wb = Workbook()
        # Eliminar hoja por defecto si existe
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    
    # Crear nueva hoja
    ws = wb.create_sheet(sheet_name)
    
    # Preparar datos para Excel
    if not results:
        # Hoja vacía con headers
        headers = ["URL", "Estado", "Código de Estado", "Error"]
        ws.append(headers)
    else:
        # Extraer todas las claves únicas de los resultados
        all_keys = set()
        for result in results:
            extracted = _extract_data_from_response(result)
            all_keys.update(extracted.keys())
        
        # Ordenar claves (URL primero, luego el resto alfabéticamente)
        sorted_keys = ["url", "success", "status_code", "error"]
        for key in sorted(all_keys):
            if key not in sorted_keys:
                sorted_keys.append(key)
        
        # Escribir headers
        ws.append(sorted_keys)
        
        # Escribir datos
        for result in results:
            extracted = _extract_data_from_response(result)
            row = [extracted.get(key, "") for key in sorted_keys]
            # Convertir objetos complejos a JSON string
            row = [
                json.dumps(cell) if isinstance(cell, (dict, list)) else str(cell)
                for cell in row
            ]
            ws.append(row)
    
    # Formatear headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo
    wb.save(excel_file)
    logger.info(f"Excel guardado en: {excel_file.absolute()}")
    
    return str(excel_file.absolute()), sheet_name


def scrape_urls(
    urls: List[str],
    profile: ProfileReadSchema,
    excel_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """Scrapea múltiples URLs y guarda los resultados en Excel.
    
    Returns:
        Diccionario con información del resultado del scraping
    """
    if not profile.tokens:
        raise ValueError("El perfil no tiene tokens asociados")
    
    if not urls:
        raise ValueError("La lista de URLs está vacía")
    
    # Validar URLs básicas
    valid_urls = []
    invalid_urls = []
    for url in urls:
        url_clean = url.strip()
        if url_clean and (url_clean.startswith("http://") or url_clean.startswith("https://")):
            valid_urls.append(url_clean)
        else:
            invalid_urls.append(url_clean)
    
    if not valid_urls:
        raise ValueError("No hay URLs válidas para scrapear")
    
    if invalid_urls:
        logger.warning(f"Se encontraron {len(invalid_urls)} URLs inválidas que serán ignoradas")
    
    results: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    
    try:
        with CrawlbaseClient(profile=profile) as client:
            for i, url in enumerate(valid_urls, 1):
                logger.info(f"Scrapeando URL {i}/{len(valid_urls)}: {url}")
                try:
                    result = scrape_url(client, url)
                    results.append(result)
                    
                    if not result["success"]:
                        errors.append({
                            "url": url,
                            "error": result.get("error", "Error desconocido"),
                        })
                except Exception as e:
                    logger.error(f"Error inesperado scrapeando {url}: {e}")
                    results.append({
                        "url": url,
                        "success": False,
                        "status_code": None,
                        "data": None,
                        "headers": {},
                        "error": str(e),
                    })
                    errors.append({
                        "url": url,
                        "error": str(e),
                    })
    except Exception as e:
        logger.error(f"Error crítico durante el scraping: {e}")
        raise ValueError(f"Error durante el scraping: {str(e)}")
    
    # Guardar en Excel
    try:
        excel_file, final_sheet_name = save_to_excel(results, excel_path, sheet_name)
        
        successful = sum(1 for r in results if r["success"])
        failed = len(results) - successful
        
        return {
            "success": True,
            "total_urls": len(valid_urls),
            "successful": successful,
            "failed": failed,
            "excel_path": excel_file,
            "sheet_name": final_sheet_name,
            "message": f"Scraping completado: {successful} exitosos, {failed} fallidos",
            "errors": errors if errors else None,
        }
    except Exception as e:
        logger.error(f"Error guardando en Excel: {e}")
        raise ValueError(f"Error guardando resultados en Excel: {str(e)}")

